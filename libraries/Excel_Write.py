from openpyxl import load_workbook
from libraries.Config import Config_Json, Key, Global
from difflib import SequenceMatcher

col_dict = dict()
for index, ascii_code in enumerate(range(65, 90)):
    col_dict[chr(ascii_code)] = index + 1


class Excel_Write:
    def __init__(self, input_json_filepath):
        self.sheet_name = str()
        self.excel_filepath = str()
        self.wb = None
        self.sheet = None
        self.json_filepath = input_json_filepath

    def load_file(self, file_path, sheet_name):
        self.sheet_name = sheet_name
        self.wb = load_workbook(file_path)
        self.sheet = self.wb[self.sheet_name]
        self.excel_filepath = file_path

    def search_name_index(self, name):
        # SequenceMatcher(None, a, b).ratio()
        config = Config_Json("json.json").read(Key.excel.key)
        col_index = 1
        for row in self.sheet.iter_rows(min_row=2, min_col=col_dict[config[Key.excel.col_name.key]],
                                        max_col=col_dict[config[Key.excel.col_name.key]]):
            col_index = col_index + 1
            for cell in row:
                if cell.value == name:
                    return col_index
        return False

    def write_data(self, name_index, value):
        self.sheet[name_index] = value

    def save_excel(self):
        try:
            self.wb.save(self.excel_filepath)
            return True
        except:
            return False


if __name__ == "__main__":
    excel = Excel_Write(Global.json_filepath)
    print(Global.json_filepath)
    excel.load_file('test.xlsx', '工作表1')
    index = excel.search_name_index("王曉薇")
    test_config = Config_Json("json.json").read(Key.excel.key)
    excel.write_data(test_config[Key.excel.col_name.key] + str(index), "asd")
    excel.save_excel()

    # # 讀取 Excel 檔案
    # wb = load_workbook('test.xlsx')
    #
    # config = Config_Json("json.json").read(Key.excel.key)
    #
    # sheet = wb['工作表1']

    # for row in sheet.iter_rows(min_row=2, min_col=col_dict[config[Key.excel.col_name.key]],
    #                            max_col=col_dict[config[Key.excel.col_name.key]]):
    #     for cell in row:
    #         print(cell.value)
